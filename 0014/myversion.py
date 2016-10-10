#! python3
# -*- coding: utf-8 -*-
"""http://blog.csdn.net/sophie2805/article/details/48143237
http://stackoverflow.com/questions/1894269/convert-string-representation-of-list-to-list-in-python
http://stackoverflow.com/questions/15004838/how-to-write-a-list-to-xlsx-using-openpyxl"""
import ast
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, colors


def standard_data():
    data = []
    output = {}
    with open('student.txt', 'r', encoding='gbk') as f:
        for line in f:
            if not line.startswith('{') and not line.startswith('}'):
                data.append(ast.literal_eval(line.split(':')[-1].replace(',\n', '').replace('\n', '')))
        for index, value in enumerate(data, start=1):
            output.setdefault(index, value)
    return output


# {1: ['张三', 150, 120, 100], 2: ['李四', 90, 99, 95], 3: ['王五', 60, 66, 68]}
# [['张三', 150, 120, 100], ['李四', 90, 99, 95], ['王五', 60, 66, 68]]


def put_data_to_xlsx():
    wb = Workbook()
    dest_filename = 'final.xlsx'
    ws = wb.create_sheet("data", 0)
    alignment = Alignment(horizontal='center', vertical='center')
    font = Font(bold=True, color=colors.RED)
    titles = ['id', '姓名', '数学', '英语', '语文']
    title_row = 1  # 标题所在行数
    for index, title in enumerate(titles, start=1):
        ws.cell(row=title_row, column=index).value = title
        ws.cell(row=title_row, column=index).font = font
        ws.cell(row=title_row, column=index).alignment = alignment

    data = standard_data()
    for key, value in enumerate(data.values(), start=1):
        ws.cell(row=key+1, column=1).value = key  # 插入#所在列的数字
        ws.cell(row=key+1, column=1).alignment = alignment
        for index, element in enumerate(value, start=2):  # 2-2, 2-3, 2-4, 2-5, 3-2, 3-3, 3-4, 3-5, 4-2, 4-3, 4-4, 4-5
            ws.cell(row=key+1, column=index).value = element
            ws.cell(row=key+1, column=index).alignment = alignment
    wb.save(dest_filename)

if __name__ == '__main__':
    print(standard_data())
    put_data_to_xlsx()
