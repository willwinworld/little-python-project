#! python3
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, colors
import re

# wb = Workbook()
#
# dest_file = 'ttt.xlsx'  # 最终要生成的文档名字
#
# ws = wb.create_sheet('student', 0)  # 创建表的名字
# # ws = wb.active  # 如果不create_sheet,就选择为默认的表sheet
#
# alignment = Alignment(horizontal='center', vertical='center')  # 设置字体是否居中
# font = Font(bold=True, color=colors.BLUE)  # 设置字体
#
# col = 'ABCDE'  # 设置列
#
# ws[col[0]+'1'] = '#'
# ws[col[1]+'1'] = '姓名'
# ws[col[2]+'1'] = '数学'
# ws[col[3]+'1'] = '英语'
# ws[col[4]+'1'] = '语文'
#
# for x in col:  # 遍历列
#     ws[x+'1'].alignment = alignment  # 将第一列第一行的属性赋值
#     ws[x+'1'].font = font
#
data = []

p = re.compile(':\[')

source_file = open('student.txt', 'r')

for line in source_file:
    if not line.startswith('{') and not line.startswith('}'):
        line = line.strip('\n], ')
        line = p.sub(',', line)
        data.append(line.split(','))

print(data)
# [['"1"', '"张三"', ' 150', ' 120', ' 100'], ['"2"', '"李四"', ' 90', ' 99', ' 95'], ['"3"', '"王五"', ' 60', ' 66', ' 68']]

source_file.close()

for i in range(len(data)):
    for j in range(5):
        ws[col[j]+str(i+2)] = data[i][j].strip('"')  # i+2是因为从第二行开始的
        print(data[i][j].strip('"'))
        ws[col[j]+str(i+2)].alignment = alignment

# wb.save(dest_file)



