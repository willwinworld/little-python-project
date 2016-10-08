#! python3
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

# wb = Workbook()
# ws = wb.create_sheet("Mysheet", 0)
# c = ws['A4']
# ws['A4'] = 4
# d = ws.cell(row=4, column=2, value=10)
# # creating cells
# for i in range(1, 101):
#     for j in range(1, 101):
#         ws.cell(row=i, column=j)
# wb.save('balances.xlsx')
wb = Workbook()
dest_filename = 'empty_book_1.xlsx'

# ws1 = wb.active
# ws1.title = "range names"
#
# for row in range(1, 40):
#     ws1.append(range(600))
#
# ws2 = wb.create_sheet(title="Pi")
#
# ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)

wb.save(filename=dest_filename)
