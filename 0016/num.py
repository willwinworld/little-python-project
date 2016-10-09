#! python3
# -*- coding: utf-8 -*-
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment


def standard_output():
    with open('numbers.txt', 'r') as f:
        c = f.read()
    data = json.loads(c)
    return data


def put_data_to_xlsx():
    wb = Workbook()
    dest_filename = 'final.xlsx'
    ws = wb.create_sheet("data", 0)
    alignment = Alignment(horizontal='center', vertical='center')
    data = standard_output()
    for outer_index, element in enumerate(data, start=1):
        for inner_index, tiny_element in enumerate(element, start=1):
            ws.cell(row=outer_index, column=inner_index).value = tiny_element
            ws.cell(row=outer_index, column=inner_index).alignment = alignment
    wb.save(dest_filename)


if __name__ == '__main__':
    put_data_to_xlsx()

