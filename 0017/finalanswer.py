#! python2
# -*- coding: utf-8 -*-
"""https://github.com/Show-Me-the-Code/python/blob/master/Liez-python-code/0017/0017.py
http://blog.csdn.net/huangxiongbiao/article/details/45974247
python2的环境"""
import json
import HTMLParser
import xml.dom.minidom as md
from openpyxl import load_workbook


def get_data_from_xlsx(filename):
    wb = load_workbook(filename)
    sheet = wb['data']
    data = []
    out = {}
    for row in sheet.iter_rows(min_row=2, max_col=5, max_row=4):
        for cell in row:
            data.append(cell.value)
    count = 0
    for _ in range(3):
        out.setdefault(data[count], data[count+1:count+5])
        count += 5
    return out


def put_data_to_xml(xlscontent):
    xmlfile = md.Document()  # 创建新的xml文件

    root = xmlfile.createElement('root')  # 创建节点
    students = xmlfile.createElement('students')  # 创建节点

    xmlfile.appendChild(root)  # 在文件中添加root节点
    root.appendChild(students)  # 在root下添加students节点

    comment = xmlfile.createComment('学生信息表 "id": [名字, 数学, 语文, 英文]')  # 创建评论
    students.appendChild(comment)  # 在students标签下添加comment

    xmlcontent = xmlfile.createTextNode(json.dumps(xlscontent).decode('unicode-escape').encode('utf-8'))  # 创建文本节点
    students.appendChild(xmlcontent)  # 在students标签下添加文本内容

    with open('students.xml', 'w') as f:
        html_parser = HTMLParser.HTMLParser()  # toprettyxml返回DOM的字符串，然后使用HTMLParser,把转义后的字符还原回来
        tranform = html_parser.unescape(xmlfile.toprettyxml().decode('utf-8'))
        f.write(tranform.encode('utf-8'))

if __name__ == '__main__':
    put_data_to_xml(get_data_from_xlsx('final.xlsx'))
