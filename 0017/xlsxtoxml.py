#! python3
# -*- coding: utf-8 -*-
"""解决在virtualenv下安装lxml的问题，http://stackoverflow.com/questions/27363697/install-lxml-into-virtualenv
http://www.lfd.uci.edu/~gohlke/pythonlibs/#lxml
把lxml的包下载下来，pip install lxml-3.4.4-cp34-none-win_amd64.whl就可以了
xhtml是实现html向xml的过渡
1、html即是超文本标记语言（Hyper Text Markup Language），是最早写网页的语言，但是由于时间早，规范不是很好，大小写混写且编码不规范；
2、xhtml即是升级版的html（Extensible Hyper Text Markup Language），对html进行了规范，编码更加严谨纯洁，也是一种过渡语言，html向xml过渡的语言；
3、xml即时可扩展标记语言（Extensible Markup Language），是一种跨平台语言，编码更自由，可以自由创建标签。
4、网页编码从html>>xhtml>>xml这个过程发展。
HTML5, like HTML 1.0, is not defined using any meta language.
It is written in English text and moves radically in opposition of the uniform requirements of an XML serialization.
 HTML5 appears to be created for usability and media delivery without regard for structure or language hierarchies.
 HTML和XML的最大区别在于：HTML是一个定型的标记语言，它用固有的标记来描述，显示网页内容。比如<H1>表示首行标题，有固定的尺寸。相对的，XML则没有固定的标记，XML不能描述网页具体的外观，内容，它只是描述内容的数据形式和结构。
这是一个质的区别：网页将数据和显示混在一起，而XML则将数据和显示分开来"""
import json
import codecs
import xml.etree.ElementTree as etree
from openpyxl import load_workbook


def get_data_from_xlsx():
    wb = load_workbook(filename='final.xlsx')
    sheet = wb['data']
    data = []
    out = {}
    for row in sheet.iter_rows(min_row=2, max_col=5, max_row=4):
        for cell in row:
            data.append(cell.value)
    # print(data)
    # print(len(data))  # key:0, 5, 10 value:[1-4], [6-9], [11-14]
    count = 0
    for _ in range(3):
        out.setdefault(data[count], data[count+1:count+5])
        count += 5
    # print(out)
    return out


def put_data_to_xml():
    data = get_data_from_xlsx()
    print(type(data))
    root = etree.Element("root")
    child = etree.SubElement(root, "students")
    # child.append(etree.Comment(u"""学生信息表 "id" : [名字, 数学, 语文, 英文]"""))
    # child.text = str(data, encoding='gbk')
    child.text = json.dumps(data)
    student_xml = etree.ElementTree(root)
    output = codecs.open("test.xml", "w", encoding="utf8")
    output.write(etree.tostring(student_xml))
    output.close()
    # student_xml.write('student.xml', encoding='utf8', xml_declaration=True)
    # print(etree.tostring(root, method="xml"))







if __name__ == '__main__':
    # get_data_from_xlsx()
    put_data_to_xml()