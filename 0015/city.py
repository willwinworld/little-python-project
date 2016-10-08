#! python3
# -*- coding: utf-8 -*-
"""http://stackoverflow.com/questions/33453188/proper-way-to-align-cell-in-python-using-openpyxl
http://stackoverflow.com/questions/11026959/python-writing-dict-to-txt-file-and-reading-dict-from-txt-file
主要学到了关于如何将txt文件中的字符串类型的数据类型转换成python的数据类型:1.通过ast模块的ast.literal_eval2.常用方法json.loads"""
import ast
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, colors


def put_data_to_xlxs():
    wb = Workbook()
    dest_filename = 'final.xlsx'
    ws = wb.create_sheet("data", 0)
    alignment = Alignment(horizontal='center', vertical='center')
    with open('city.txt', 'r', encoding='gbk') as f:
        s = f.read()
    content = json.loads(s)
    for key, value in content.items():
        ws.cell(row=int(key), column=1).value = int(key)
        ws.cell(row=int(key), column=1).alignment = alignment
        ws.cell(row=int(key), column=2).value = value
        ws.cell(row=int(key), column=2).alignment = alignment
    wb.save(dest_filename)

if __name__ == '__main__':
    put_data_to_xlxs()

