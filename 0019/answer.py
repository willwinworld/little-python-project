#! python2
# -*- coding: utf-8 -*-
import json
import HTMLParser
import xml.dom.minidom as md
from openpyxl import load_workbook


def get_data_from_xlsx(filename):
    wb = load_workbook(filename)
    sheet = wb['data']
    data = []
    out = []
    for row in sheet.iter_rows(min_row=1, max_col=3, max_row=3):
        for cell in row:
            data.append(cell.value)
    # print(data)
    count = 0
    for _ in range(3):
        out.append(data[count:count+3])
        count += 3
    return out


def put_data_to_xml(xlscontent):
    xmlfile = md.Document()  # 创建新的xml文件

    root = xmlfile.createElement('root')  # 创建节点
    numbers = xmlfile.createElement('numbers')  # 创建节点

    xmlfile.appendChild(root)  # 在文件中添加root节点
    root.appendChild(numbers)  # 在root下添加students节点

    comment = xmlfile.createComment('numbers')  # 创建评论
    numbers.appendChild(comment)  # 在students标签下添加comment

    xmlcontent = xmlfile.createTextNode(json.dumps(xlscontent, sort_keys=True, indent=4).decode('unicode-escape').encode('utf-8'))  # 创建文本节点
    numbers.appendChild(xmlcontent)  # 在students标签下添加文本内容

    with open('numbers.xml', 'w') as f:
        html_parser = HTMLParser.HTMLParser()  # toprettyxml返回DOM的字符串，然后使用HTMLParser,把转义后的字符还原回来
        tranform = html_parser.unescape(xmlfile.toprettyxml().decode('utf-8'))
        f.write(tranform.encode('utf-8'))

if __name__ == '__main__':
    put_data_to_xml(get_data_from_xlsx('final.xlsx'))
    # get_data_from_xlsx('final.xlsx')